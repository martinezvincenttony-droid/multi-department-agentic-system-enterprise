#!/usr/bin/env python3
"""Excel Power Updater - Advanced Excel manipulation engine.

Usage:
    python excel_updater.py --input <file> --operation <op> [--options ...]

Operations:
    update_cells    - Update cells based on criteria
    cross_reference - Cross-reference two files
    add_column      - Add a new column with formula or values
    format          - Apply professional formatting
    merge           - Merge multiple sheets/files
    filter_extract  - Extract filtered subset
    pivot_summary   - Create pivot-style summary
    full_format     - Apply complete professional formatting
"""

import argparse
import json
import os
import shutil
from copy import copy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter


# ── Styling Presets ──────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def backup_file(filepath):
    """Create a backup before modifying."""
    backup = filepath.replace(".xlsx", "_backup.xlsx")
    shutil.copy2(filepath, backup)
    print(f"Backup created: {backup}")
    return backup


def apply_professional_formatting(ws, header_row=1):
    """Apply professional formatting to a worksheet."""
    # Format headers
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Format data rows with alternating colors and borders
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Auto-fit column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    print("Professional formatting applied.")


def cross_reference(file1, file2, key_col1, key_col2, pull_cols, output_file):
    """Cross-reference two Excel files and pull matching data."""
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    ws1 = wb1.active
    ws2 = wb2.active

    # Build lookup from file2
    headers2 = [cell.value for cell in ws2[1]]
    key_idx2 = headers2.index(key_col2) if isinstance(key_col2, str) else key_col2 - 1
    pull_indices = []
    for pc in pull_cols:
        if isinstance(pc, str):
            pull_indices.append(headers2.index(pc))
        else:
            pull_indices.append(pc - 1)

    lookup = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        key = row[key_idx2]
        if key:
            lookup[str(key).strip().lower()] = [row[i] for i in pull_indices]

    # Add new columns to file1
    headers1 = [cell.value for cell in ws1[1]]
    key_idx1 = headers1.index(key_col1) if isinstance(key_col1, str) else key_col1 - 1
    start_col = ws1.max_column + 1

    for i, pc in enumerate(pull_cols):
        col_name = pc if isinstance(pc, str) else headers2[pc - 1]
        ws1.cell(row=1, column=start_col + i, value=f"{col_name} (from ref)")

    matched = 0
    for row_num in range(2, ws1.max_row + 1):
        key = ws1.cell(row=row_num, column=key_idx1 + 1).value
        if key and str(key).strip().lower() in lookup:
            values = lookup[str(key).strip().lower()]
            for i, val in enumerate(values):
                ws1.cell(row=row_num, column=start_col + i, value=val)
            matched += 1

    apply_professional_formatting(ws1)
    wb1.save(output_file)
    print(f"Cross-reference complete. {matched} rows matched. Saved to {output_file}")


def update_cells_by_criteria(filepath, sheet_name, criteria_col, criteria_val, update_col, new_value, output_file):
    """Update cells where criteria column matches a value."""
    backup_file(filepath)
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [cell.value for cell in ws[1]]
    crit_idx = headers.index(criteria_col) + 1 if isinstance(criteria_col, str) else criteria_col
    upd_idx = headers.index(update_col) + 1 if isinstance(update_col, str) else update_col

    updated = 0
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=crit_idx).value
        if str(cell_val).strip().lower() == str(criteria_val).strip().lower():
            ws.cell(row=row, column=upd_idx, value=new_value)
            updated += 1

    wb.save(output_file)
    print(f"Updated {updated} cells. Saved to {output_file}")


def add_column_with_formula(filepath, sheet_name, col_name, formula_template, output_file):
    """Add a new column with an Excel formula."""
    backup_file(filepath)
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=col_name)

    for row in range(2, ws.max_row + 1):
        formula = formula_template.replace("{ROW}", str(row))
        ws.cell(row=row, column=new_col, value=formula)

    apply_professional_formatting(ws)
    wb.save(output_file)
    print(f"Column '{col_name}' added. Saved to {output_file}")


def merge_workbooks(file_list, output_file, same_structure=True):
    """Merge multiple workbooks into one."""
    combined = Workbook()
    combined.remove(combined.active)

    if same_structure:
        merged_ws = combined.create_sheet("Merged Data")
        header_written = False
        for f in file_list:
            wb = load_workbook(f)
            ws = wb.active
            for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_num == 1 and header_written:
                    continue
                merged_ws.append(list(row))
                if row_num == 1:
                    header_written = True
        apply_professional_formatting(merged_ws)
    else:
        for f in file_list:
            wb = load_workbook(f)
            for ws in wb.worksheets:
                new_ws = combined.create_sheet(title=ws.title[:31])
                for row in ws.iter_rows():
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.fill = copy(cell.fill)
                            new_cell.border = copy(cell.border)
                            new_cell.alignment = copy(cell.alignment)

    combined.save(output_file)
    print(f"Merged {len(file_list)} files. Saved to {output_file}")


def create_pivot_summary(filepath, group_col, value_col, agg_func, output_file):
    """Create a pivot-style summary sheet."""
    wb = load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    grp_idx = headers.index(group_col) if isinstance(group_col, str) else group_col - 1
    val_idx = headers.index(value_col) if isinstance(value_col, str) else value_col - 1

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[grp_idx]
        val = row[val_idx]
        if key not in data:
            data[key] = []
        if val is not None:
            try:
                data[key].append(float(val))
            except (ValueError, TypeError):
                pass

    # Create summary sheet
    summary_ws = wb.create_sheet("Pivot Summary")
    summary_ws.append([group_col, f"{agg_func.upper()}({value_col})", "Count"])

    for key, values in sorted(data.items(), key=lambda x: str(x[0])):
        if agg_func == "sum":
            result = sum(values)
        elif agg_func == "avg":
            result = sum(values) / len(values) if values else 0
        elif agg_func == "max":
            result = max(values) if values else 0
        elif agg_func == "min":
            result = min(values) if values else 0
        else:
            result = len(values)
        summary_ws.append([key, round(result, 2), len(values)])

    apply_professional_formatting(summary_ws)
    wb.save(output_file)
    print(f"Pivot summary created. Saved to {output_file}")


if __name__ == "__main__":
    print("Excel Power Updater ready. Import functions or run with arguments.")
