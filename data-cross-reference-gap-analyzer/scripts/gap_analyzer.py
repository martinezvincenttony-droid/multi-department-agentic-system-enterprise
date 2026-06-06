#!/usr/bin/env python3
"""Data Cross-Reference & Gap Analyzer - Multi-file reconciliation engine."""

import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ── Color Coding ───────────────────────────────────────────────
CRITICAL_FILL = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
HIGH_FILL = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
LOW_FILL = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
COMPLIANT_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SEVERITY_FILLS = {
    "CRITICAL": CRITICAL_FILL,
    "HIGH": HIGH_FILL,
    "MEDIUM": MEDIUM_FILL,
    "LOW": LOW_FILL,
    "COMPLIANT": COMPLIANT_FILL,
}


def load_sheet_data(filepath, sheet_name=None):
    """Load an Excel sheet and return headers + rows as list of dicts."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    data = []
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        record = dict(zip(headers, row))
        data.append(record)
    return headers, data


def normalize_key(value):
    """Normalize a key value for matching."""
    if value is None:
        return ""
    return str(value).strip().lower()


def cross_reference_files(file1, file2, key_col1, key_col2, compare_cols=None,
                          sheet1=None, sheet2=None):
    """Cross-reference two files on key columns and identify gaps."""
    headers1, data1 = load_sheet_data(file1, sheet1)
    headers2, data2 = load_sheet_data(file2, sheet2)

    # Build lookup from file2
    lookup2 = {}
    for record in data2:
        key = normalize_key(record.get(key_col2))
        if key:
            lookup2[key] = record

    results = {
        "matched": [],
        "in_file1_only": [],
        "in_file2_only": [],
        "mismatches": [],
        "total_file1": len(data1),
        "total_file2": len(data2),
    }

    matched_keys = set()
    for record1 in data1:
        key = normalize_key(record1.get(key_col1))
        if not key:
            continue
        if key in lookup2:
            matched_keys.add(key)
            record2 = lookup2[key]
            # Check for mismatches in compare columns
            if compare_cols:
                for col in compare_cols:
                    val1 = record1.get(col)
                    val2 = record2.get(col)
                    if normalize_key(val1) != normalize_key(val2):
                        results["mismatches"].append({
                            "key": record1.get(key_col1),
                            "column": col,
                            "file1_value": val1,
                            "file2_value": val2,
                        })
            results["matched"].append({"key": record1.get(key_col1), "file1": record1, "file2": record2})
        else:
            results["in_file1_only"].append(record1)

    # Find records only in file2
    for record2 in data2:
        key = normalize_key(record2.get(key_col2))
        if key and key not in matched_keys:
            results["in_file2_only"].append(record2)

    return results


def generate_gap_report(results, output_file, file1_name="File 1", file2_name="File 2",
                        key_col_name="Key"):
    """Generate a comprehensive gap analysis Excel report."""
    wb = Workbook()

    # ── Summary Dashboard ──
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"

    total = results["total_file1"] + results["total_file2"]
    matched = len(results["matched"])
    gaps_f1 = len(results["in_file1_only"])
    gaps_f2 = len(results["in_file2_only"])
    mismatches = len(results["mismatches"])
    compliance_pct = round((matched / max(results["total_file1"], 1)) * 100, 1)

    summary_data = [
        ["Metric", "Value"],
        ["Total Records (File 1)", results["total_file1"]],
        ["Total Records (File 2)", results["total_file2"]],
        ["Matched Records", matched],
        [f"Only in {file1_name}", gaps_f1],
        [f"Only in {file2_name}", gaps_f2],
        ["Value Mismatches", mismatches],
        ["Match Rate", f"{compliance_pct}%"],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            elif col_idx == 1:
                cell.font = BOLD_FONT
            else:
                cell.font = NORMAL_FONT
                # Color code compliance
                if row_idx == 8:  # Match Rate row
                    if compliance_pct >= 90:
                        cell.fill = COMPLIANT_FILL
                    elif compliance_pct >= 70:
                        cell.fill = MEDIUM_FILL
                    else:
                        cell.fill = CRITICAL_FILL
                        cell.font = WHITE_FONT

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20

    # ── Gaps: Only in File 1 ──
    if results["in_file1_only"]:
        ws_gaps1 = wb.create_sheet(f"Only in {file1_name[:20]}")
        if results["in_file1_only"]:
            headers = list(results["in_file1_only"][0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws_gaps1.cell(row=1, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN
            for row_idx, record in enumerate(results["in_file1_only"], 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_gaps1.cell(row=row_idx, column=col_idx, value=record.get(header))
                    cell.border = THIN_BORDER
                    cell.font = NORMAL_FONT
                    if row_idx % 2 == 0:
                        cell.fill = ALT_ROW_FILL

    # ── Gaps: Only in File 2 ──
    if results["in_file2_only"]:
        ws_gaps2 = wb.create_sheet(f"Only in {file2_name[:20]}")
        if results["in_file2_only"]:
            headers = list(results["in_file2_only"][0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws_gaps2.cell(row=1, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN
            for row_idx, record in enumerate(results["in_file2_only"], 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_gaps2.cell(row=row_idx, column=col_idx, value=record.get(header))
                    cell.border = THIN_BORDER
                    cell.font = NORMAL_FONT
                    if row_idx % 2 == 0:
                        cell.fill = ALT_ROW_FILL

    # ── Mismatches ──
    if results["mismatches"]:
        ws_mismatch = wb.create_sheet("Mismatches")
        mismatch_headers = ["Key", "Column", f"{file1_name} Value", f"{file2_name} Value"]
        for col_idx, header in enumerate(mismatch_headers, 1):
            cell = ws_mismatch.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        for row_idx, mm in enumerate(results["mismatches"], 2):
            values = [mm["key"], mm["column"], mm["file1_value"], mm["file2_value"]]
            for col_idx, value in enumerate(values, 1):
                cell = ws_mismatch.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.font = NORMAL_FONT
                if row_idx % 2 == 0:
                    cell.fill = ALT_ROW_FILL

    # Auto-width all sheets
    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    wb.save(output_file)
    print(f"Gap analysis report generated: {output_file}")
    print(f"\nSummary:")
    print(f"  Matched: {matched}")
    print(f"  Only in {file1_name}: {gaps_f1}")
    print(f"  Only in {file2_name}: {gaps_f2}")
    print(f"  Mismatches: {mismatches}")
    print(f"  Match Rate: {compliance_pct}%")
    return output_file


def training_matrix_analysis(training_file, pa_file, output_file,
                              training_sheet=None, pa_sheet=None,
                              employee_col="Employee", skill_col="Skill",
                              status_col="Status", required_col="Required"):
    """Specialized analysis for Training Matrix vs PA Matrix."""
    _, training_data = load_sheet_data(training_file, training_sheet)
    _, pa_data = load_sheet_data(pa_file, pa_sheet)

    # Build training lookup: {employee: {skill: status}}
    training_lookup = defaultdict(dict)
    for record in training_data:
        emp = normalize_key(record.get(employee_col, ""))
        skill = normalize_key(record.get(skill_col, ""))
        status = record.get(status_col, "Unknown")
        if emp and skill:
            training_lookup[emp][skill] = status

    # Build requirements: {role/area: [required_skills]}
    requirements = defaultdict(list)
    for record in pa_data:
        role = record.get("Role", record.get("Area", "General"))
        skill = record.get(skill_col, record.get(required_col, ""))
        if skill:
            requirements[normalize_key(str(role))].append({
                "skill": skill,
                "skill_key": normalize_key(str(skill)),
            })

    # Identify gaps
    gaps = []
    compliant = []
    for emp, skills in training_lookup.items():
        for role, req_skills in requirements.items():
            for req in req_skills:
                if req["skill_key"] in skills:
                    status = skills[req["skill_key"]]
                    if normalize_key(str(status)) in ["certified", "complete", "yes", "trained", "current", "x"]:
                        compliant.append({"employee": emp, "skill": req["skill"], "status": status})
                    else:
                        gaps.append({
                            "employee": emp,
                            "skill": req["skill"],
                            "current_status": status,
                            "severity": "MEDIUM",
                        })
                else:
                    gaps.append({
                        "employee": emp,
                        "skill": req["skill"],
                        "current_status": "NOT FOUND",
                        "severity": "CRITICAL",
                    })

    print(f"Training Analysis Complete:")
    print(f"  Employees analyzed: {len(training_lookup)}")
    print(f"  Compliant items: {len(compliant)}")
    print(f"  Gaps found: {len(gaps)}")
    print(f"  Critical gaps: {sum(1 for g in gaps if g['severity'] == 'CRITICAL')}")

    return gaps, compliant


if __name__ == "__main__":
    print("Data Cross-Reference & Gap Analyzer ready.")
