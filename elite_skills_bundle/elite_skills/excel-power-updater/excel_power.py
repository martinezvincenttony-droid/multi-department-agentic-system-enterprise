"""
Excel Power Updater — core operations module.
Uses openpyxl (preinstalled) for all Excel manipulation.
"""
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Formatting ----------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_formatting(ws):
    """Apply professional formatting to a worksheet."""
    if ws.max_row == 0:
        return
    # Header styling
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    # Alternating row fill + borders
    for r in range(2, ws.max_row + 1):
        for cell in ws[r]:
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL
    # Auto column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, 50))
        ws.column_dimensions[col_letter].width = max_len
    # Freeze header
    ws.freeze_panes = "A2"


# ---------- Operations ----------
def update_cells(src_path, sheet, match_col, match_value, update_col, new_value, out_path=None):
    """Update cells in update_col where match_col == match_value."""
    wb = load_workbook(src_path)
    ws = wb[sheet]
    headers = {c.value: c.column for c in ws[1]}
    mc, uc = headers[match_col], headers[update_col]
    updates = 0
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, mc).value) == str(match_value):
            ws.cell(r, uc).value = new_value
            updates += 1
    apply_formatting(ws)
    out_path = out_path or _stamped(src_path)
    wb.save(out_path)
    return {"updates": updates, "output": out_path}


def add_formula_column(src_path, sheet, new_header, formula_template, out_path=None):
    """
    Add a formula column. formula_template uses {row} placeholder.
    Example: '=B{row}+C{row}'  or  '=VLOOKUP(A{row},Sheet2!A:B,2,FALSE)'
    """
    wb = load_workbook(src_path)
    ws = wb[sheet]
    new_col = ws.max_column + 1
    ws.cell(1, new_col).value = new_header
    for r in range(2, ws.max_row + 1):
        ws.cell(r, new_col).value = formula_template.format(row=r)
    apply_formatting(ws)
    out_path = out_path or _stamped(src_path)
    wb.save(out_path)
    return {"column_added": new_header, "output": out_path}


def cross_reference(primary_path, lookup_path, key_col, pull_cols, out_path=None):
    """VLOOKUP-style merge: enrich primary file with columns from lookup file."""
    primary = load_workbook(primary_path)
    lookup = load_workbook(lookup_path)
    pws, lws = primary.active, lookup.active

    # Build lookup dict
    l_headers = {c.value: c.column for c in lws[1]}
    key_idx = l_headers[key_col]
    lookup_map = {}
    for r in range(2, lws.max_row + 1):
        k = lws.cell(r, key_idx).value
        if k is not None:
            lookup_map[str(k)] = {col: lws.cell(r, l_headers[col]).value for col in pull_cols if col in l_headers}

    p_headers = {c.value: c.column for c in pws[1]}
    p_key = p_headers[key_col]

    # Append new columns
    start_col = pws.max_column + 1
    for i, col in enumerate(pull_cols):
        pws.cell(1, start_col + i).value = col
    matched = 0
    for r in range(2, pws.max_row + 1):
        k = str(pws.cell(r, p_key).value)
        if k in lookup_map:
            matched += 1
            for i, col in enumerate(pull_cols):
                pws.cell(r, start_col + i).value = lookup_map[k].get(col)

    apply_formatting(pws)
    out_path = out_path or _stamped(primary_path)
    primary.save(out_path)
    return {"matched_rows": matched, "total_rows": pws.max_row - 1, "output": out_path}


def merge_workbooks(file_list, out_path):
    """Merge multiple workbooks into a single workbook (one sheet per file)."""
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    for f in file_list:
        src = load_workbook(f, data_only=True)
        for sname in src.sheetnames:
            sws = src[sname]
            new_name = f"{os.path.basename(f)[:20]}_{sname}"[:31]
            new_ws = out_wb.create_sheet(new_name)
            for row in sws.iter_rows(values_only=True):
                new_ws.append(row)
            apply_formatting(new_ws)
    out_wb.save(out_path)
    return {"sheets": len(out_wb.sheetnames), "output": out_path}


def update_training_matrix(src_path, employee, process, status, cert_date=None, out_path=None):
    """Specialized: mark an employee as certified on a process with date."""
    wb = load_workbook(src_path)
    ws = wb.active
    headers = {c.value: c.column for c in ws[1]}
    if process not in headers:
        return {"error": f"Process '{process}' not found in matrix."}
    pcol = headers[process]
    name_col = 1  # assume first column is employee name
    cert_date = cert_date or datetime.today().strftime("%Y-%m-%d")
    val = f"{status} ({cert_date})" if status else None
    updated = False
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, name_col).value).strip().lower() == employee.strip().lower():
            ws.cell(r, pcol).value = val
            updated = True
            break
    apply_formatting(ws)
    out_path = out_path or _stamped(src_path)
    wb.save(out_path)
    return {"updated": updated, "employee": employee, "process": process, "output": out_path}


def _stamped(path):
    base, ext = os.path.splitext(path)
    return f"{base}_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"


import os  # placed here to keep top-section clean
