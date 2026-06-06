"""
Data Cross-Reference & Gap Analyzer — multi-file reconciliation + Excel reports.
"""
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEVERITY_COLORS = {
    "CRITICAL":  "C00000",
    "HIGH":      "ED7D31",
    "MEDIUM":    "FFC000",
    "LOW":       "9DC3E6",
    "COMPLIANT": "70AD47",
}
SEVERITY_FONT = {"CRITICAL": "FFFFFF", "HIGH": "FFFFFF", "MEDIUM": "000000", "LOW": "000000", "COMPLIANT": "FFFFFF"}


def cross_reference_files(primary_path, reference_path, key_column, primary_sheet=None, reference_sheet=None):
    """Compare two files on a key column. Returns rows present, missing, or mismatched."""
    p_wb = load_workbook(primary_path, data_only=True)
    r_wb = load_workbook(reference_path, data_only=True)
    pws = p_wb[primary_sheet] if primary_sheet else p_wb.active
    rws = r_wb[reference_sheet] if reference_sheet else r_wb.active

    p_data = _sheet_to_dicts(pws)
    r_data = _sheet_to_dicts(rws)

    p_keys = {str(row.get(key_column, "")).strip() for row in p_data if row.get(key_column)}
    r_keys = {str(row.get(key_column, "")).strip() for row in r_data if row.get(key_column)}

    return {
        "primary_count": len(p_data),
        "reference_count": len(r_data),
        "missing_in_primary": sorted(r_keys - p_keys),
        "missing_in_reference": sorted(p_keys - r_keys),
        "common": sorted(p_keys & r_keys),
        "primary_data": p_data,
        "reference_data": r_data,
    }


def _sheet_to_dicts(ws):
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers)) if headers[i]}
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows


def classify_gaps(cross_ref_result, critical_processes=None, expiry_days=365):
    """
    Classify each gap by severity.
    critical_processes: list of process names that are compliance-critical
    """
    critical_processes = critical_processes or []
    gaps = []
    today = datetime.today()

    # Missing from primary = people/processes not yet trained/certified
    for missing in cross_ref_result["missing_in_primary"]:
        is_critical = any(c.lower() in missing.lower() for c in critical_processes)
        gaps.append({
            "item": missing,
            "type": "Missing Certification",
            "severity": "CRITICAL" if is_critical else "HIGH",
            "description": f"{missing} present in reference file but missing from primary",
            "action": "Schedule training and update primary record",
        })

    # Check for expired/old certs in primary data
    for row in cross_ref_result.get("primary_data", []):
        for col, val in row.items():
            if val and isinstance(val, datetime):
                age_days = (today - val).days
                if age_days > expiry_days:
                    gaps.append({
                        "item": f"{row.get('Name', row.get('Employee', '?'))} - {col}",
                        "type": "Expired Certification",
                        "severity": "HIGH" if age_days < expiry_days * 1.5 else "CRITICAL",
                        "description": f"Last certified {age_days} days ago",
                        "action": "Renew certification within 30 days",
                    })

    # Single-point-of-failure detection
    process_counts = {}
    for row in cross_ref_result.get("primary_data", []):
        for col, val in row.items():
            if val and col not in ("Name", "Employee", "ID"):
                process_counts.setdefault(col, 0)
                process_counts[col] += 1
    for process, count in process_counts.items():
        if count == 1:
            gaps.append({
                "item": process,
                "type": "Single Point of Failure",
                "severity": "CRITICAL" if any(c.lower() in process.lower() for c in critical_processes) else "MEDIUM",
                "description": f"Only 1 person certified for {process}",
                "action": "Cross-train at least 1 additional person",
            })

    return gaps


def build_remediation_plan(gaps):
    """Convert gaps into prioritized action items with timelines."""
    severity_to_days = {"CRITICAL": 7, "HIGH": 30, "MEDIUM": 90, "LOW": 180}
    today = datetime.today()
    plan = []
    for i, g in enumerate(sorted(gaps, key=lambda x: list(SEVERITY_COLORS).index(x["severity"])), 1):
        days = severity_to_days.get(g["severity"], 60)
        plan.append({
            "priority": f"P{1 if g['severity']=='CRITICAL' else 2 if g['severity']=='HIGH' else 3}",
            "item": g["item"],
            "severity": g["severity"],
            "action": g["action"],
            "owner": "TBD",
            "due_date": (today + timedelta(days=days)).strftime("%Y-%m-%d"),
            "status": "Open",
        })
    return plan


def export_gap_report(gaps, remediation_plan, out_path, source_summary=None):
    """Build a 3-sheet color-coded Excel gap analysis report."""
    wb = Workbook()
    wb.remove(wb.active)

    # ----- Sheet 1: Summary Dashboard -----
    s1 = wb.create_sheet("Summary Dashboard")
    _write_summary(s1, gaps, source_summary)

    # ----- Sheet 2: Detailed Gaps -----
    s2 = wb.create_sheet("Detailed Gaps")
    _write_detailed_gaps(s2, gaps)

    # ----- Sheet 3: Remediation Plan -----
    s3 = wb.create_sheet("Remediation Plan")
    _write_remediation(s3, remediation_plan)

    wb.save(out_path)
    return out_path


def _write_summary(ws, gaps, source_summary):
    ws["A1"] = "GAP ANALYSIS — EXECUTIVE SUMMARY"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A4"] = f"Total Gaps Identified: {len(gaps)}"

    # Severity breakdown
    ws["A6"] = "Severity Breakdown"
    ws["A6"].font = Font(bold=True, size=12)
    counts = {sev: sum(1 for g in gaps if g["severity"] == sev) for sev in SEVERITY_COLORS}
    row = 7
    ws.cell(row, 1).value = "Severity"; ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2).value = "Count"; ws.cell(row, 2).font = Font(bold=True)
    for sev, count in counts.items():
        row += 1
        ws.cell(row, 1).value = sev
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=SEVERITY_COLORS[sev])
        ws.cell(row, 1).font = Font(color=SEVERITY_FONT[sev], bold=True)
        ws.cell(row, 2).value = count

    # Top 3 priorities
    ws.cell(row + 2, 1).value = "Top 3 Priority Actions"
    ws.cell(row + 2, 1).font = Font(bold=True, size=12)
    top = sorted(gaps, key=lambda x: list(SEVERITY_COLORS).index(x["severity"]))[:3]
    for i, g in enumerate(top, 1):
        ws.cell(row + 2 + i, 1).value = f"{i}. [{g['severity']}] {g['item']}"
        ws.cell(row + 2 + i, 2).value = g["action"]

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 30


def _write_detailed_gaps(ws, gaps):
    headers = ["Item", "Type", "Severity", "Description", "Recommended Action"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center")
    for r, g in enumerate(gaps, 2):
        ws.cell(r, 1).value = g["item"]
        ws.cell(r, 2).value = g["type"]
        ws.cell(r, 3).value = g["severity"]
        ws.cell(r, 3).fill = PatternFill("solid", fgColor=SEVERITY_COLORS[g["severity"]])
        ws.cell(r, 3).font = Font(bold=True, color=SEVERITY_FONT[g["severity"]])
        ws.cell(r, 4).value = g["description"]
        ws.cell(r, 5).value = g["action"]
    for col, w in zip(range(1, 6), [30, 25, 14, 50, 40]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def _write_remediation(ws, plan):
    headers = ["Priority", "Item", "Severity", "Action", "Owner", "Due Date", "Status"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center")
    for r, p in enumerate(plan, 2):
        ws.cell(r, 1).value = p["priority"]
        ws.cell(r, 2).value = p["item"]
        ws.cell(r, 3).value = p["severity"]
        ws.cell(r, 3).fill = PatternFill("solid", fgColor=SEVERITY_COLORS[p["severity"]])
        ws.cell(r, 3).font = Font(bold=True, color=SEVERITY_FONT[p["severity"]])
        ws.cell(r, 4).value = p["action"]
        ws.cell(r, 5).value = p["owner"]
        ws.cell(r, 6).value = p["due_date"]
        ws.cell(r, 7).value = p["status"]
    for col, w in zip(range(1, 8), [10, 30, 14, 40, 18, 14, 12]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
